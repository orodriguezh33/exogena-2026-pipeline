# src/exportar.py
"""Escritura de informes a CSV con formato adecuado para análisis.

Los montos se redondean a pesos ENTEROS (como los pide la DIAN). Al no tener
decimales, el número se interpreta igual en Excel (es-CO) y en Google Sheets,
sin el problema del separador decimal (punto vs coma). Se guarda con BOM
UTF-8 para que Excel respete tildes y ñ.
"""
import re

import numpy as np
import pandas as pd


def a_entero(s: pd.Series) -> pd.Series:
    """Redondea a entero medio-hacia-arriba conservando el signo (NC negativas)."""
    f = pd.to_numeric(s, errors="coerce")
    return (np.sign(f) * np.floor(np.abs(f) + 0.5)).astype("Int64")


def _limpiar_texto(v):
    """Quita espacios sobrantes (extremos y dobles internos) de los textos."""
    if isinstance(v, str):
        return re.sub(r"\s+", " ", v).strip()
    return v


def escribir_csv(df: pd.DataFrame, ruta, cols_moneda: tuple = ()) -> None:
    """Escribe el DataFrame a CSV: dinero a pesos enteros y textos sin espacios sobrantes."""
    df = df.copy()
    for c in df.columns:
        if c not in cols_moneda and (df[c].dtype == object or pd.api.types.is_string_dtype(df[c])):
            df[c] = df[c].map(_limpiar_texto)
    for c in cols_moneda:
        if c in df.columns:
            df[c] = a_entero(df[c])
    df.to_csv(ruta, index=False, encoding="utf-8-sig")


def escribir_excel(
    df: pd.DataFrame,
    ruta,
    cols_texto: tuple = (),
    cols_moneda: tuple = (),
) -> None:
    """
    Escribe a .xlsx. Las cols_texto (códigos con ceros a la izquierda, NIT, CUFE)
    se guardan como TEXTO real → Excel/Sheets muestran '001', no 1. Las cols_moneda
    como número entero con formato de miles. Encabezado en negrita y fila congelada.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    df = df.copy()
    for c in df.columns:
        if c not in cols_moneda and (df[c].dtype == object or pd.api.types.is_string_dtype(df[c])):
            df[c] = df[c].map(_limpiar_texto)
    for c in cols_moneda:
        if c in df.columns:
            df[c] = a_entero(df[c])

    wb = Workbook()
    ws = wb.active
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    def _clean(v):
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        if isinstance(v, np.integer):
            return int(v)
        if isinstance(v, np.floating):
            return float(v)
        return v

    for fila in df.itertuples(index=False, name=None):
        ws.append([_clean(v) for v in fila])

    idx = {name: i + 1 for i, name in enumerate(df.columns)}
    ult = len(df) + 1
    for c in cols_texto:
        if c in idx:
            letra = get_column_letter(idx[c])
            for r in range(2, ult + 1):
                ws[f"{letra}{r}"].number_format = "@"
    for c in cols_moneda:
        if c in idx:
            letra = get_column_letter(idx[c])
            for r in range(2, ult + 1):
                ws[f"{letra}{r}"].number_format = "#,##0"

    for name, i in idx.items():
        letra = get_column_letter(i)
        ws.column_dimensions[letra].width = (
            42 if any(k in name for k in ("nombre", "direccion")) else
            18 if (name in cols_moneda or name == "ultimo_cufe") else 16
        )
    ws.freeze_panes = "A2"
    if len(df.columns):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{ult}"
    wb.save(ruta)
